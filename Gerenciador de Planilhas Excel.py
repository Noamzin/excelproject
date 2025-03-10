import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
import pandas as pd
from tkinter import *
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import numbers
from datetime import datetime

FORMATACOES_POR_PLANILHA = {}


def converter_valor(valor, formato):
    """Converte o valor inserido para o tipo correto com base na formatação"""
    try:
        if formato == 'R$ #,##0.00':
            valor = valor.replace('R$', '').replace(' ', '').replace(',', '.')
            return float(valor)  # Garante que seja float
        elif formato == numbers.FORMAT_PERCENTAGE:
            return float(valor.replace('%', '').replace(',', '.')) / 100
        elif formato == 'dd/mm/yyyy':
            return datetime.strptime(valor, '%d/%m/%Y')
        else:
            return valor
    except ValueError:
        messagebox.showinfo("ERRO",
                            f"Não foi possível converter '{valor}' para o formato esperado. Mantido como texto.")
        return valor


def selecionar_arquivo():
    """Função para selecionar o arquivo diretamente"""
    return filedialog.askopenfilename(filetypes=[("Arquivos Excel", "*.xlsx")])



def criar_planilha():
    """Cria uma planilha do excel vazia"""
    output_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Arquivos Excel", "*.xlsx")],
        title="Salvar Relatório Como",
        initialfile="Planilha"  # Nome padrão sugerido
    )
    if not output_path:  # Se o usuário cancelar a seleção
        messagebox.showinfo("Cancelado", "Salvamento do relatório foi cancelado.")
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Planilha Vazia"
    wb.save(output_path)
    messagebox.showinfo("Sucesso", f"Planilha {output_path} criada com sucesso!")


def obter_formatacao_escolhida(opcao):
    """Retorna a formatação escolhida"""
    formatos = {
        '1': numbers.FORMAT_GENERAL,
        '2': 'R$ #,##0.00',
        '3': numbers.FORMAT_PERCENTAGE,
        '4': 'dd/mm/yyyy'
    }
    return formatos.get(opcao, numbers.FORMAT_GENERAL)


def carregar_formatacoes_planilha(caminho_arquivo):
    """Carrega as formatações de uma planilha que foi importada"""
    global FORMATACOES_POR_PLANILHA
    try:
        wb = load_workbook(caminho_arquivo)
        ws = wb.active
        formatacoes = {}
        colunas = {ws.cell(1, i + 1).value: i + 1 for i in range(ws.max_column) if ws.cell(1, i + 1).value}

        for col_name, col_idx in colunas.items():
            celula = ws.cell(row=2, column=col_idx)
            formato = celula.number_format if celula.number_format else numbers.FORMAT_GENERAL
            # Mapeia formatos de moeda do Excel para 'R$ #,##0.00'
            if any(f in formato for f in ['#,##0.00', 'R$', '"R$"']):
                formato = 'R$ #,##0.00'
            formatacoes[col_name] = formato

        max_linhas = ws.max_row
        formatacoes["MAX_LINHAS"] = max_linhas
        FORMATACOES_POR_PLANILHA[caminho_arquivo] = formatacoes
        return formatacoes
    except Exception as e:
        messagebox.showinfo("ERRO", f"Erro ao carregar formatações: {e}")
        return {}


def adicionar_colunas():
    """Adiciona colunas numa tabela já existente"""
    try:
        global FORMATACOES_POR_PLANILHA
        caminho_arquivo = selecionar_arquivo()
        if not caminho_arquivo:
            messagebox.showinfo("ERRO", "Arquivo não encontrado!")
            return

        df = pd.read_excel(caminho_arquivo)
        novas_colunas = simpledialog.askstring("Adicionar Colunas",
                                               "Digite os nomes das novas colunas separados por vírgula:")
        if not novas_colunas:
            return

        novas_colunas = [col.strip() for col in novas_colunas.split(',')]
        for col in novas_colunas:
            df[col] = None

        df.to_excel(caminho_arquivo, index=False)
        formatacoes = FORMATACOES_POR_PLANILHA.get(caminho_arquivo, {})
        for col in novas_colunas:
            opcao = simpledialog.askstring("Formatar Colunas\n",
                                           f"Escolha a formatação para: {col}\n"
                                           "1 - Geral\n"
                                           "2 - Contábil (R$ #,##0.00)\n"
                                           "3 - Percentual\n"
                                           "4 - Data")
            formatacoes[col] = obter_formatacao_escolhida(opcao)
        FORMATACOES_POR_PLANILHA[caminho_arquivo] = formatacoes

        df.to_excel(caminho_arquivo, index=False)
        aplicar_formatacoes(caminho_arquivo, formatacoes)
        messagebox.showinfo("Sucesso",f"Novas colunas adicionadas e formatadas até a linha {FORMATACOES_POR_PLANILHA[caminho_arquivo]['MAX_LINHAS']}")
    except Exception as e:
        messagebox.showinfo("ERRO", f"Erro ao adicionar colunas: {e}")


def aplicar_formatacoes(output_path, formatacoes):
    """Aplica as formatações às colunas da planilha"""
    try:
        wb = load_workbook(output_path)
        ws = wb.active
        colunas_idx = {ws.cell(1, i + 1).value: i + 1 for i in range(ws.max_column)}

        # Se já há um número salvo, reutiliza. Senão, pede ao usuário.
        if output_path in FORMATACOES_POR_PLANILHA and "MAX_LINHAS" in FORMATACOES_POR_PLANILHA[output_path]:
            MAX_LINHAS = FORMATACOES_POR_PLANILHA[output_path]["MAX_LINHAS"]
        else:
            MAX_LINHAS = simpledialog.askinteger("Número Máximo de Linhas", "Digite quantas linhas da planilha você deseja formatar: ")
            FORMATACOES_POR_PLANILHA[output_path]["MAX_LINHAS"] = MAX_LINHAS  # Armazena a escolha do usuário

        for col, formato in formatacoes.items():
            if col in colunas_idx:
                idx = colunas_idx[col]
                for row in range(2, MAX_LINHAS + 1):
                    celula = ws.cell(row=row, column=idx)
                    celula.number_format = formato

        wb.save(output_path)
    except Exception as e:
        messagebox.showinfo("ERRO", f"Houve um problema ao aplicar as formatações: {e}")


def adicionar_dados():
    """Função que adiciona dados a uma planilha já existente"""
    global FORMATACOES_POR_PLANILHA
    try:
        caminho_arquivo = selecionar_arquivo()
        if not caminho_arquivo:
            messagebox.showinfo("ERRO", "Arquivo não encontrado!")
            return

        formatacoes = carregar_formatacoes_planilha(caminho_arquivo)
        df = pd.read_excel(caminho_arquivo)
        messagebox.showinfo("Colunas", f"Colunas: {list(df.columns)}")
        messagebox.showinfo("Formatações Carregadas", f"Formatos: {formatacoes}")

        if df.empty or df.isna().all().all():
            df = pd.DataFrame(columns=df.columns)
            primeiras_linhas = True
        else:
            primeiras_linhas = False

        while True:
            nova_linha = {}
            for coluna in df.columns:
                formato = formatacoes.get(coluna, numbers.FORMAT_GENERAL)
                valor = simpledialog.askstring("Adicionar Dados", f"Digite o valor para '{coluna}' ({formato}):")
                if valor is not None:
                    nova_linha[coluna] = converter_valor(valor, formato)
                else:
                    nova_linha[coluna] = None

            if primeiras_linhas:
                df = pd.DataFrame([nova_linha])
                primeiras_linhas = False
            else:
                df = pd.concat([df, pd.DataFrame([nova_linha])], ignore_index=True)

            continuar = simpledialog.askstring("CONTINUAR ?", "Deseja adicionar outra linha (S/N)? ").strip().upper()
            if continuar != 'S':
                break

        # Converte colunas de moeda para float explicitamente antes de salvar
        for coluna in df.columns:
            if formatacoes.get(coluna) == 'R$ #,##0.00':
                df[coluna] = pd.to_numeric(df[coluna], errors='coerce')

        df.to_excel(caminho_arquivo, index=False)
        aplicar_formatacoes(caminho_arquivo, formatacoes)
        messagebox.showinfo("Sucesso", "Dados adicionados com êxito.")
    except Exception as e:
        messagebox.showinfo("ERRO", f"Houve um problema ao adicionar os dados: {e}")


def gerar_relatorios():
    """Gera relatórios personalizados a partir de uma planilha com parâmetros definidos pelo usuário"""
    try:
        caminho_arquivo = selecionar_arquivo()
        if not caminho_arquivo:
            messagebox.showinfo("ERRO", "Arquivo não encontrado.")
            return

        df = pd.read_excel(caminho_arquivo)
        messagebox.showinfo("Colunas", f"Colunas disponíveis na planilha: {list(df.columns)}")

        coluna_agrupamento = simpledialog.askstring("Agrupamento", "Digite a coluna para agrupar os dados (ex.: PRODUTO, REMETENTE): ").strip()
        if coluna_agrupamento not in df.columns:
            messagebox.showinfo("ERRO", f"Coluna {coluna_agrupamento} não encontrada na planilha!")
            return
        messagebox.showinfo("Funções disponíveis", "Funções de agregação disponíveis: \n"
                                        f"1. Soma\n"
                                        f"2. Média\n"
                                        f"3. Contagem\n"
                                        f"4. Máximo\n"
                                        f"5. Mínimo")
        opcao_agregacao = simpledialog.askstring("Agregação","Escolha a função de agregação (1-5): ").strip()
        agregacao_map = {
            '1': 'sum',
            '2': 'mean',
            '3': 'count',
            '4': 'max',
            '5': 'min'
        }
        agregacao = agregacao_map.get(opcao_agregacao, 'sum')

        colunas_numericas = []
        if agregacao != 'count':
            colunas_numericas = simpledialog.askstring("Escolha as colunas","Digite as colunas numéricas para agregar (separadas por vírgula, ex.: FRETE, PESO LÍQUIDO): ").split(',')
            colunas_numericas = [col.strip() for col in colunas_numericas]
            colunas_invalidas = [col for col in colunas_numericas if col not in df.columns]
            if colunas_invalidas:
                messagebox.showinfo("ERRO", f"Colunas inválidas: {colunas_invalidas}!")
                return

            # Gera o relatório com base na agregação escolhida
        if agregacao == 'count':
                # Cria uma nova coluna 'Contagem' com o número de linhas por grupo
            relatorio = df.groupby(coluna_agrupamento).size().reset_index(name='Contagem')
        else:
                # Aplica a agregação nas colunas numéricas selecionadas
            relatorio = df.groupby(coluna_agrupamento)[colunas_numericas].agg(agregacao).reset_index()

            # Ordenação opcional
        ordenar = simpledialog.askstring("Deseja ordenar ?","Deseja ordenar alguma coluna? (S/N):" ).upper().strip()[0]
        if ordenar == 'S':
            coluna_ordenacao = simpledialog.askstring("Ordenar coluna", f"Digite a coluna para ordenar (ex.: {', '.join(relatorio.columns)}): ")
            if coluna_ordenacao in relatorio.columns:
                ordem = simpledialog.askstring("Crescente ou Decrescente ?", "Ordem crescente (S) ou decrescente (N)? (S/N): ").upper().strip()[0]
                relatorio = relatorio.sort_values(by=coluna_ordenacao, ascending=(ordem == 'S'))
            else:
                    messagebox.showinfo("ERRO", f"Coluna {coluna_ordenacao} não encontrada no relatório. Ordem padrão mantida.")

        output_path_novo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Arquivos Excel", "*.xlsx")],
            title="Salvar Relatório Como",
            initialfile="Relatório"  # Nome padrão sugerido
        )
        if not output_path_novo:  # Se o usuário cancelar a seleção
            messagebox.showinfo("Cancelado", "Salvamento do relatório foi cancelado.")
            return

            # Salva o relatório
        with pd.ExcelWriter(output_path_novo, engine='openpyxl') as writer:
            relatorio.to_excel(writer, sheet_name='Relatório Personalizado', index=False)

            # Aplica formatações
        wb_novo = load_workbook(output_path_novo)
        ws_novo = wb_novo['Relatório Personalizado']
        colunas_idx_novo = {ws_novo.cell(1, i + 1).value: i + 1 for i in range(ws_novo.max_column)}

            # Lê as formatações da planilha original
        wb_original = load_workbook(caminho_arquivo)
        ws_original = wb_original.active
        formatacoes_originais = {}
        colunas_idx_original = {ws_original.cell(1, i + 1).value: i + 1 for i in range(ws_original.max_column)}
        for col_name, col_idx in colunas_idx_original.items():
            cell = ws_original.cell(row=2, column=col_idx)
            formatacoes_originais[col_name] = cell.number_format if cell.value is not None else numbers.FORMAT_GENERAL

            # Aplica as formatações às colunas do relatório
        for col_name, formato in formatacoes_originais.items():
            if col_name in colunas_idx_novo:
                idx = colunas_idx_novo[col_name]
                for row in range(2, ws_novo.max_row + 1):
                    celula = ws_novo.cell(row=row, column=idx)
                    celula.number_format = formato

            # Formatação específica para a coluna 'Contagem'
        if agregacao == 'count' and 'Contagem' in colunas_idx_novo:
            idx_contagem = colunas_idx_novo['Contagem']
            for row in range(2, ws_novo.max_row + 1):
                celula = ws_novo.cell(row=row, column=idx_contagem)
                celula.number_format = numbers.FORMAT_NUMBER  # Número geral para 'Contagem'

        wb_novo.save(output_path_novo)
        messagebox.showinfo("SUCESSO", f"Relatório salvo em '{output_path_novo}' com formatações aplicadas!")
    except Exception as e:
        messagebox.showinfo("ERRO", f"Erro ao gerar relatório: {e}")



def main():
    """Função principal"""
    root = tk.Tk()
    root.title("Gerenciador de Planilhas Excel")
    root.geometry("400x300")
    root.configure(bg="#f0f0f0")

    # Define o ícone da janela
    try:
        # Caminho relativo ao diretório do script ou .exe
        if hasattr(sys, '_MEIPASS'):  # Caso esteja rodando como .exe
            icon_path = os.path.join(sys._MEIPASS, "icone.ico")
        else:  # Caso esteja rodando como script
            icon_path = os.path.join(os.path.dirname(__file__), "icone.ico")
        root.iconbitmap(icon_path)
    except Exception as e:
        print(f"Erro ao carregar o ícone: {e}")  # Debug, pode ser removido depois

    Label(root, text="Gerenciador de Planilhas", font=("Arial", 16, "bold"), bg="#f0f0f0").pack(pady=20)
    btn_style = {"font": ("Arial", 12), "width": 30, "bg": "#4CAF50", "fg": "white", "activebackground": "#45a049"}
    btn_criar = Button(root, text="Criar Planilha", command=criar_planilha, **btn_style)
    btn_criar.pack(pady=5)
    btn_adicionar_colunas = Button(root, text="Adicionar Colunas", command=adicionar_colunas, **btn_style)
    btn_adicionar_colunas.pack(pady=5)
    btn_adicionar_dados = Button(root, text="Adicionar Dados", command=adicionar_dados, **btn_style)
    btn_adicionar_dados.pack(pady=5)
    btn_relatorio = Button(root, text="Gerar Relatório", command=gerar_relatorios, **btn_style)
    btn_relatorio.pack(pady=5)
    btn_sair = Button(root, text="Sair", command=root.quit, font=("Arial", 12), width=30, bg="#f44336", fg="white",
                      activebackground="#d32f2f")
    btn_sair.pack(pady=5)

    root.mainloop()

if __name__ == "__main__":
    import sys
    main()
